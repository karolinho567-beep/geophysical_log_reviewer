"""Write the review deliverable: one Excel row per log, with the evidence attached.

The workbook is the point of the whole tool. It is built so that a person can
confirm or overturn every machine decision without opening a single TIFF:

* **A** the log file,
* **B** whether a stamp was found,
* **C** a link that opens the page *at the stamp* -- a rendered neighbourhood,
  because a hyperlink cannot position a TIFF viewer at a pixel and these pages run
  up to 110 ft long,
* **D** the stamp itself, cropped, stood upright, embedded in the cell.

Rows with no detection still get a page thumbnail in **D**, so a negative is
something a reviewer can confirm rather than take on faith.
"""
from __future__ import annotations

import os
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .detection import PageResult

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
VERDICT_FILL = {
    "YES": PatternFill("solid", fgColor="C6EFCE"),
    "MAYBE": PatternFill("solid", fgColor="FFEB9C"),
    "NO": PatternFill("solid", fgColor="F2F2F2"),
}

#: (header, width). Order is the user-facing contract: file, verdict, link, image.
COLUMNS: list[tuple[str, int]] = [
    ("Log file", 42),
    ("Stamp?", 9),
    ("Open at stamp", 20),
    ("Stamp image", 34),
    ("Reviewed", 11),
    ("Score", 8),
    ("Ring seen", 10),
    ("Page X (in)", 11),
    ("Page Y (in)", 11),
    ("Diam (in)", 10),
    ("Rotation", 10),
    ("API-14", 16),
    ("DPI", 7),
    ("Page W x H (in)", 16),
    ("Other candidates", 17),
    ("Notes / warnings", 40),
]

ROW_HEIGHT = 132   # points, sized to the embedded crop
IMAGE_PX = 168     # crop is rendered to this and anchored in column D


@dataclass
class RowAssets:
    """Rendered files backing one row, as paths relative to the workbook."""

    crop_rel: str | None = None
    context_rel: str | None = None


def verdict_for(page: PageResult) -> str:
    if page.hits:
        return "YES"
    if page.uncertain:
        return "MAYBE"
    return "NO"


def write_workbook(
    path: str,
    pages: list[PageResult],
    assets: dict[str, RowAssets],
    title: str = "Date stamps",
) -> str:
    """Build the workbook. ``assets`` is keyed by ``PageResult.source_file``."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    for col, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30

    review = DataValidation(type="list", formula1='"OK,CHECK,FAIL"', allow_blank=True)
    review.error = "Choose OK, CHECK or FAIL"
    ws.add_data_validation(review)

    workbook_dir = os.path.dirname(os.path.abspath(path))

    for i, page in enumerate(sorted(pages, key=lambda p: p.source_file), start=2):
        asset = assets.get(page.source_file, RowAssets())
        verdict = verdict_for(page)
        best = page.best
        name = os.path.basename(page.source_file)

        ws.cell(row=i, column=1, value=name).alignment = Alignment(vertical="center", wrap_text=True)

        vcell = ws.cell(row=i, column=2, value=verdict)
        vcell.fill = VERDICT_FILL[verdict]
        vcell.font = Font(bold=True)
        vcell.alignment = Alignment(horizontal="center", vertical="center")

        # C -- the link. Relative, so the workbook and its images move together.
        link_cell = ws.cell(row=i, column=3)
        if asset.context_rel:
            link_cell.value = "open at stamp" if best else "open page"
            link_cell.hyperlink = asset.context_rel.replace("\\", "/")
            link_cell.font = Font(color="0563C1", underline="single")
        else:
            link_cell.value = "-"
        link_cell.alignment = Alignment(horizontal="center", vertical="center")

        # D -- the picture, anchored into the cell.
        if asset.crop_rel:
            abs_crop = os.path.join(workbook_dir, asset.crop_rel)
            if os.path.exists(abs_crop):
                picture = XLImage(abs_crop)
                scale = min(1.0, IMAGE_PX / max(picture.width, picture.height))
                picture.width = int(picture.width * scale)
                picture.height = int(picture.height * scale)
                picture.anchor = f"D{i}"
                ws.add_image(picture)

        rcell = ws.cell(row=i, column=5)
        review.add(rcell)
        rcell.alignment = Alignment(horizontal="center", vertical="center")

        others = max(0, len([d for d in page.detections if d.decision != "miss"]) - 1)
        warn = "; ".join(page.warnings)
        if page.dpi_source == "default":
            warn = "; ".join(filter(None, [warn, "resolution tag missing - assumed 400 dpi"]))
        if best is not None and best.attrs.get("clipped"):
            warn = "; ".join(filter(None, [warn, "stamp runs off the page edge"]))

        values = [
            (6, None if best is None else round(best.score, 3)),
            (7, None if best is None else round(best.evidence.get("raw_ring_completeness", 0), 2)),
            (8, None if best is None else best.center_in_x),
            (9, None if best is None else best.center_in_y),
            (10, None if best is None else round(best.radius_in * 2, 2)),
            (11, None if best is None or best.attrs.get("upright_ccw_deg") is None
                 else f"{int(best.attrs['upright_ccw_deg'])}°"),
            (12, page.api14),
            (13, round(page.dpi)),
            (14, f"{page.width / page.dpi:.1f} x {page.height / page.dpi:.1f}"),
            (15, others or None),
            (16, warn or None),
        ]
        for col, value in values:
            ws.cell(row=i, column=col, value=value).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=(col == 16))

        ws.row_dimensions[i].height = ROW_HEIGHT

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(pages) + 1}"

    _add_summary(wb, pages)
    os.makedirs(workbook_dir, exist_ok=True)
    wb.save(path)
    return path


def _add_summary(wb: Workbook, pages: list[PageResult]) -> None:
    """A second sheet with the counts, so the workbook reports its own totals."""
    ws = wb.create_sheet("Summary")
    counts = {"YES": 0, "MAYBE": 0, "NO": 0}
    for page in pages:
        counts[verdict_for(page)] += 1
    rows = [
        ("Logs examined", len(pages)),
        ("Stamp found (YES)", counts["YES"]),
        ("Needs a look (MAYBE)", counts["MAYBE"]),
        ("No stamp found (NO)", counts["NO"]),
        ("", ""),
        ("Total scan time (s)", round(sum(p.seconds for p in pages), 1)),
        ("Pages with warnings", sum(1 for p in pages if p.warnings)),
    ]
    for r, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=label).font = Font(bold=(r <= 4))
        ws.cell(row=r, column=2, value=value)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14


def write_csv(path: str, pages: list[PageResult]) -> str:
    """Flat one-row-per-detection table, for joining to other datasets on API-14."""
    import csv

    rows = []
    for page in pages:
        base = {
            "source_file": os.path.basename(page.source_file),
            "api14": page.api14,
            "dpi": page.dpi,
            "page_w_in": round(page.width / page.dpi, 2),
            "page_h_in": round(page.height / page.dpi, 2),
            "verdict": verdict_for(page),
        }
        if not page.detections:
            rows.append(base)
            continue
        for det in page.detections:
            row = dict(base)
            row.update(det.to_row())
            rows.append(row)

    fields: list[str] = []
    for row in rows:
        for key in row:
            if key not in fields:
                fields.append(key)
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return path
