"""Flexible identifier-list/manifest import for LogReview.

The review UI deliberately delegates parsing and path resolution to this module
so the behavior can be tested without Tk.  Source files are never modified.
"""
from __future__ import annotations

import csv
import datetime as _dt
import os
import re
import threading
from collections import OrderedDict, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook

from . import pages


IDENTIFIER_ALIASES = {
    "api", "api14", "apinumber", "api_number", "uwi", "wellid", "well_id",
    "wellapi", "well_api", "logapi", "log_api",
}
PATH_ALIASES = {
    "tifpath", "tif_path", "tiffpath", "tiff_path", "filepath", "file_path",
    "imagepath", "image_path", "logpath", "log_path",
}
LATITUDE_ALIASES = {
    "lat", "latitude", "welllat", "well_lat", "welllatitude", "well_latitude",
}
LONGITUDE_ALIASES = {
    "long", "lon", "lng", "longitude", "welllong", "well_long", "welllon",
    "well_lon", "welllongitude", "well_longitude",
}
DEPTH_ALIASES = {
    "depth", "depth1", "depth_1", "welldepth", "well_depth", "totaldepth",
    "total_depth", "depthft", "depth_ft", "depthfeet", "depth_feet",
}
LOG_TOP_DEPTH_ALIASES = {
    "depth1", "depth_1", "logtop", "log_top", "logtopdepth",
    "log_top_depth", "topdepth", "top_depth", "startdepth", "start_depth",
    "intervaltop", "interval_top", "fromdepth", "from_depth", "depthfrom",
    "depth_from", "availablefrom", "available_from",
}
LOG_BOTTOM_DEPTH_ALIASES = {
    "depth2", "depth_2", "logbottom", "log_bottom", "logbottomdepth",
    "log_bottom_depth", "bottomdepth", "bottom_depth", "enddepth", "end_depth",
    "intervalbottom", "interval_bottom", "todepth", "to_depth", "depthto",
    "depth_to", "availableto", "available_to",
}
PLACEHOLDERS = {"ihs assoc image"}
AUDIT_COLUMNS = [
    "import_row", "source_identifier", "match_identifier", "source_tif_path",
    "resolved_tif_path", "well_loaded", "row_image_loaded", "load_status",
    "resolution_source", "detail",
]


class ManifestError(ValueError):
    """A selected source cannot be interpreted safely."""


class ImportCancelled(ManifestError):
    """The user cancelled image resolution before it completed."""


class ColumnSelectionRequired(ManifestError):
    """The UI must ask which columns carry identifiers and paths."""

    def __init__(self, table: "SourceTable", id_candidates: list[str],
                 path_candidates: list[str], latitude_candidates: list[str],
                 longitude_candidates: list[str], depth_candidates: list[str],
                 log_top_depth_candidates: list[str],
                 log_bottom_depth_candidates: list[str]):
        super().__init__("Select the identifier and optional TIFF-path columns.")
        self.table = table
        self.id_candidates = id_candidates
        self.path_candidates = path_candidates
        self.latitude_candidates = latitude_candidates
        self.longitude_candidates = longitude_candidates
        self.depth_candidates = depth_candidates
        self.log_top_depth_candidates = log_top_depth_candidates
        self.log_bottom_depth_candidates = log_bottom_depth_candidates


@dataclass(frozen=True)
class SourceTable:
    source_path: str
    headers: tuple[str, ...]
    rows: tuple[OrderedDict[str, str], ...]
    identifier_column: str | None = None
    path_column: str | None = None
    latitude_column: str | None = None
    longitude_column: str | None = None
    depth_column: str | None = None
    log_top_depth_column: str | None = None
    log_bottom_depth_column: str | None = None

    @property
    def has_paths(self) -> bool:
        return bool(self.path_column)


@dataclass
class AuditRow:
    original: OrderedDict[str, str]
    import_row: int
    source_identifier: str
    match_identifier: str
    source_tif_path: str
    resolved_tif_path: str = ""
    well_loaded: bool = False
    row_image_loaded: bool = False
    load_status: str = ""
    resolution_source: str = ""
    detail: str = ""

    def as_dict(self) -> OrderedDict[str, object]:
        out: OrderedDict[str, object] = OrderedDict(self.original)
        for name in AUDIT_COLUMNS:
            out[name] = getattr(self, name)
        return out


@dataclass
class ImportStats:
    input_rows: int = 0
    valid_unique_identifiers: int = 0
    invalid_unique_identifiers: int = 0
    unique_tiff_candidates: int = 0
    loaded_tiffs: int = 0
    loaded_from_folder: int = 0
    recovered_from_paths: int = 0
    duplicate_references: int = 0
    loaded_identifiers: int = 0
    partially_loaded_identifiers: int = 0
    unrecovered_identifiers: int = 0
    placeholder_rows: int = 0
    placeholder_covered: int = 0
    placeholder_unresolved: int = 0
    eligible_unique_identifiers: int = 0
    depth_filtered_rows: int = 0
    depth_filtered_identifiers: int = 0
    invalid_depth_rows: int = 0
    depth_filter_description: str = ""
    unavailable_network_shares: int = 0
    network_paths_skipped: int = 0

    def summary(self) -> str:
        depth_lines = ""
        if self.depth_filter_description:
            depth_lines = (
                f"\nDepth filter: {self.depth_filter_description}\n"
                f"Rows excluded by depth filter: {self.depth_filtered_rows}\n"
                f"Rows with invalid depth excluded: {self.invalid_depth_rows}\n"
                f"Identifiers eligible after depth filter: "
                f"{self.eligible_unique_identifiers}\n"
                f"Identifiers fully excluded by depth filter: "
                f"{self.depth_filtered_identifiers}"
            )
        network_lines = ""
        if self.unavailable_network_shares:
            network_lines = (
                f"\nNetwork shares unavailable or timed out: "
                f"{self.unavailable_network_shares}\n"
                f"Network paths skipped: {self.network_paths_skipped}"
            )
        return (
            f"Input rows: {self.input_rows}\n"
            f"Valid unique identifiers: {self.valid_unique_identifiers}\n"
            f"Invalid unique identifiers: {self.invalid_unique_identifiers}\n"
            f"Unique TIFF candidates: {self.unique_tiff_candidates}\n"
            f"Successfully loaded TIFFs: {self.loaded_tiffs}\n"
            f"Loaded from selected folder: {self.loaded_from_folder}\n"
            f"Recovered from listed paths: {self.recovered_from_paths}\n"
            f"Duplicate references skipped: {self.duplicate_references}\n"
            f"Unique identifiers loaded: {self.loaded_identifiers}\n"
            f"Partially loaded identifiers: {self.partially_loaded_identifiers}\n"
            f"Identifiers not recovered: {self.unrecovered_identifiers}\n"
            f"IHS placeholder rows: {self.placeholder_rows} "
            f"({self.placeholder_covered} covered, "
            f"{self.placeholder_unresolved} unresolved)"
            f"{depth_lines}{network_lines}"
        )


@dataclass
class ImportResult:
    source: SourceTable
    paths: list[str]
    path_identifiers: dict[str, str]
    path_locations: dict[str, dict[str, str]]
    loaded_identifiers: list[str]
    audit_rows: list[AuditRow]
    stats: ImportStats


@dataclass(frozen=True)
class DepthFilter:
    """Strict row filter applied to a detected numeric depth column."""

    column: str
    comparison: str
    threshold: float
    include_blank: bool = False

    def __post_init__(self) -> None:
        if self.comparison not in {"less_than", "greater_than"}:
            raise ValueError("Depth comparison must be less_than or greater_than.")

    @property
    def description(self) -> str:
        symbol = "<" if self.comparison == "less_than" else ">"
        blanks = "including" if self.include_blank else "excluding"
        return f"{self.column} {symbol} {self.threshold:g} ft, {blanks} blank depths"


@dataclass
class _LogicalCandidate:
    identifier: str
    basename: str
    row_indices: list[int] = field(default_factory=list)
    listed_paths: list[str] = field(default_factory=list)
    folder_path: str = ""
    resolved_path: str = ""
    resolution_source: str = ""
    failure_status: str = "path_missing"
    failure_detail: str = ""


def _header_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


def _alias_keys(values: Iterable[str]) -> set[str]:
    return {_header_key(value) for value in values}


_ID_KEYS = _alias_keys(IDENTIFIER_ALIASES)
_PATH_KEYS = _alias_keys(PATH_ALIASES)
_LAT_KEYS = _alias_keys(LATITUDE_ALIASES)
_LONG_KEYS = _alias_keys(LONGITUDE_ALIASES)
_DEPTH_KEYS = _alias_keys(DEPTH_ALIASES)
_LOG_TOP_DEPTH_KEYS = _alias_keys(LOG_TOP_DEPTH_ALIASES)
_LOG_BOTTOM_DEPTH_KEYS = _alias_keys(LOG_BOTTOM_DEPTH_ALIASES)


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _unique_headers(values: Iterable[object]) -> tuple[str, ...]:
    out: list[str] = []
    used: set[str] = set()
    for index, value in enumerate(values, start=1):
        base = _cell_text(value) or f"column_{index}"
        name = base
        suffix = 2
        while name.casefold() in used:
            name = f"{base}_{suffix}"
            suffix += 1
        used.add(name.casefold())
        out.append(name)
    return tuple(out)


def _read_text_list(path: str) -> SourceTable:
    with open(path, encoding="utf-8-sig") as handle:
        values = [line.strip() for line in handle.read().splitlines() if line.strip()]
    rows = tuple(OrderedDict([("input_value", value)]) for value in values)
    return SourceTable(os.path.abspath(path), ("input_value",), rows,
                       identifier_column="input_value")


def _read_delimited(path: str) -> SourceTable:
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(65536)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except csv.Error:
            dialect = csv.excel_tab if Path(path).suffix.lower() == ".tsv" else csv.excel
        reader = csv.reader(handle, dialect)
        try:
            header_values = next(reader)
        except StopIteration as exc:
            raise ManifestError("The selected file is empty.") from exc
        headers = _unique_headers(header_values)
        rows: list[OrderedDict[str, str]] = []
        for values in reader:
            padded = list(values) + [""] * max(0, len(headers) - len(values))
            row = OrderedDict((name, _cell_text(padded[i]))
                              for i, name in enumerate(headers))
            if any(row.values()):
                rows.append(row)
    return SourceTable(os.path.abspath(path), headers, tuple(rows))


def _read_xlsx(path: str) -> SourceTable:
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        fallback: SourceTable | None = None
        for sheet in book.worksheets:
            values = sheet.iter_rows(values_only=True)
            try:
                header_values = next(values)
            except StopIteration:
                continue
            headers = _unique_headers(header_values)
            rows: list[OrderedDict[str, str]] = []
            for cells in values:
                padded = list(cells) + [None] * max(0, len(headers) - len(cells))
                row = OrderedDict((name, _cell_text(padded[i]))
                                  for i, name in enumerate(headers))
                if any(row.values()):
                    rows.append(row)
            table = SourceTable(os.path.abspath(path), headers, tuple(rows))
            if fallback is None:
                fallback = table
            if any(_header_key(h) in _ID_KEYS for h in headers):
                return table
        if fallback is not None:
            return fallback
    finally:
        book.close()
    raise ManifestError("The selected workbook contains no non-empty worksheets.")


def read_source(path: str) -> SourceTable:
    """Read a supported source without interpreting its columns."""
    suffix = Path(path).suffix.lower()
    try:
        if suffix == ".txt":
            return _read_text_list(path)
        if suffix in {".csv", ".tsv"}:
            return _read_delimited(path)
        if suffix == ".xlsx":
            return _read_xlsx(path)
    except (OSError, UnicodeError) as exc:
        raise ManifestError(str(exc)) from exc
    raise ManifestError("Supported list formats are .txt, .csv, .tsv, and .xlsx.")


def select_columns(table: SourceTable, identifier_column: str | None = None,
                   path_column: str | None = None,
                   latitude_column: str | None = None,
                   longitude_column: str | None = None,
                   depth_column: str | None = None,
                   log_top_depth_column: str | None = None,
                   log_bottom_depth_column: str | None = None) -> SourceTable:
    """Attach selected columns, auto-detecting unambiguous aliases."""
    headers = list(table.headers)
    id_candidates = [h for h in headers if _header_key(h) in _ID_KEYS]
    path_candidates = [h for h in headers if _header_key(h) in _PATH_KEYS]
    latitude_candidates = [h for h in headers if _header_key(h) in _LAT_KEYS]
    longitude_candidates = [h for h in headers if _header_key(h) in _LONG_KEYS]
    depth_candidates = [h for h in headers if _header_key(h) in _DEPTH_KEYS]
    log_top_depth_candidates = [
        h for h in headers if _header_key(h) in _LOG_TOP_DEPTH_KEYS
    ]
    log_bottom_depth_candidates = [
        h for h in headers if _header_key(h) in _LOG_BOTTOM_DEPTH_KEYS
    ]
    if identifier_column is None:
        if table.identifier_column:
            identifier_column = table.identifier_column
        elif len(id_candidates) == 1:
            identifier_column = id_candidates[0]
    if path_column is None and len(path_candidates) == 1:
        path_column = path_candidates[0]
    if latitude_column is None and len(latitude_candidates) == 1:
        latitude_column = latitude_candidates[0]
    if longitude_column is None and len(longitude_candidates) == 1:
        longitude_column = longitude_candidates[0]
    if depth_column is None and len(depth_candidates) == 1:
        depth_column = depth_candidates[0]
    if log_top_depth_column is None and len(log_top_depth_candidates) == 1:
        log_top_depth_column = log_top_depth_candidates[0]
    if log_bottom_depth_column is None and len(log_bottom_depth_candidates) == 1:
        log_bottom_depth_column = log_bottom_depth_candidates[0]
    needs_id = identifier_column not in headers
    ambiguous_path = len(path_candidates) > 1 and path_column not in headers
    ambiguous_lat = len(latitude_candidates) > 1 and latitude_column not in headers
    ambiguous_long = len(longitude_candidates) > 1 and longitude_column not in headers
    ambiguous_depth = len(depth_candidates) > 1 and depth_column not in headers
    ambiguous_log_top = (
        len(log_top_depth_candidates) > 1 and log_top_depth_column not in headers
    )
    ambiguous_log_bottom = (
        len(log_bottom_depth_candidates) > 1 and log_bottom_depth_column not in headers
    )
    if (needs_id or ambiguous_path or ambiguous_lat or ambiguous_long
            or ambiguous_depth or ambiguous_log_top or ambiguous_log_bottom):
        raise ColumnSelectionRequired(
            table, id_candidates, path_candidates, latitude_candidates,
            longitude_candidates, depth_candidates, log_top_depth_candidates,
            log_bottom_depth_candidates,
        )
    for label, selected in (
        ("TIFF-path", path_column), ("latitude", latitude_column),
        ("longitude", longitude_column), ("depth", depth_column),
        ("log top depth", log_top_depth_column),
        ("log bottom depth", log_bottom_depth_column),
    ):
        if selected and selected not in headers:
            raise ManifestError(f"Unknown {label} column: {selected}")
    return SourceTable(table.source_path, table.headers, table.rows,
                       identifier_column, path_column, latitude_column,
                       longitude_column, depth_column, log_top_depth_column,
                       log_bottom_depth_column)


def load_manifest(path: str, identifier_column: str | None = None,
                  path_column: str | None = None,
                  latitude_column: str | None = None,
                  longitude_column: str | None = None,
                  depth_column: str | None = None,
                  log_top_depth_column: str | None = None,
                  log_bottom_depth_column: str | None = None) -> SourceTable:
    return select_columns(read_source(path), identifier_column, path_column,
                          latitude_column, longitude_column, depth_column,
                          log_top_depth_column, log_bottom_depth_column)


def normalize_identifier(value: str) -> tuple[str, tuple[str, ...]]:
    """Return stored identifier and filename-match variants."""
    identifier = str(value or "").strip()
    if not identifier.isdigit() or not 1 <= len(identifier) <= 14:
        return "", ()
    variants = [identifier]
    if len(identifier) == 12:
        variants.append(identifier + "00")
    return identifier, tuple(variants)


def _name_matches_identifier(name: str, variants: Iterable[str]) -> bool:
    lower = os.path.basename(name).casefold()
    for value in variants:
        if lower.startswith(value.casefold()):
            if len(lower) == len(value) or not lower[len(value)].isdigit():
                return True
    return False


def _probe(path: str, probe_fn: Callable[[str], object]) -> tuple[bool, str, str]:
    if not path.lower().endswith(pages.IMAGE_EXTS):
        return False, "unsupported_image_path", "Unsupported image extension."
    if not os.path.isfile(path):
        return False, "path_missing", "File does not exist or is not accessible."
    try:
        probe_fn(path)
    except Exception as exc:
        return False, "unreadable_image", str(exc)
    return True, "", ""


def _unc_share_root(path: str) -> str:
    r"""Return ``\\server\share`` without touching the referenced filesystem."""
    value = str(path or "").replace("/", "\\")
    if not value.startswith("\\\\"):
        return ""
    parts = [part for part in value[2:].split("\\") if part]
    if len(parts) < 2:
        return ""
    return f"\\\\{parts[0]}\\{parts[1]}"


def _network_root_status(
    root: str, timeout: float, check_fn: Callable[[str], bool]
) -> tuple[bool, str]:
    """Bound one UNC-share availability check with a daemon helper thread."""
    finished = threading.Event()
    state: dict[str, object] = {}

    def check() -> None:
        try:
            state["available"] = bool(check_fn(root))
        except Exception as exc:  # network/provider errors are ordinary failures
            state["error"] = str(exc)
        finally:
            finished.set()

    threading.Thread(
        target=check, daemon=True, name="logreview-network-preflight"
    ).start()
    if not finished.wait(max(0.05, float(timeout))):
        return False, f"Network share did not respond within {timeout:g} seconds."
    if state.get("available"):
        return True, ""
    return False, str(state.get("error") or "Network share is not accessible.")


def resolve_manifest(table: SourceTable, folder_images: Iterable[str] = (),
                     probe_fn: Callable[[str], object] = pages.probe,
                     depth_filter: DepthFilter | None = None,
                     progress_fn: Callable[[int, int, str], None] | None = None,
                     cancel_event: threading.Event | None = None,
                     max_workers: int = 8,
                     network_timeout: float = 4.0,
                     network_check: Callable[[str], bool] = os.path.isdir) -> ImportResult:
    """Resolve a manifest into readable logical images and row-level audit data."""
    if not table.identifier_column:
        raise ManifestError("No identifier column was selected.")
    if depth_filter and depth_filter.column not in table.headers:
        raise ManifestError(f"Unknown depth column: {depth_filter.column}")
    folder_paths = [os.path.abspath(path) for path in folder_images]
    folder_by_name = {os.path.basename(path).casefold(): path for path in folder_paths}
    audit: list[AuditRow] = []
    valid_order: list[str] = []
    all_valid_order: list[str] = []
    variants_by_id: dict[str, tuple[str, ...]] = {}
    all_valid_ids: set[str] = set()
    eligible_ids: set[str] = set()
    invalid_values: set[str] = set()
    rows_by_id: dict[str, list[int]] = defaultdict(list)
    depth_filtered_rows = 0
    invalid_depth_rows = 0

    for row_number, original in enumerate(table.rows, start=2):
        raw_id = original.get(table.identifier_column, "")
        identifier, variants = normalize_identifier(raw_id)
        raw_path = original.get(table.path_column, "") if table.path_column else ""
        item = AuditRow(OrderedDict(original), row_number, raw_id, variants[-1] if variants else "",
                        raw_path)
        audit.append(item)
        if not identifier:
            item.load_status = "invalid_identifier"
            item.detail = "Identifier must contain 1 to 14 digits."
            invalid_values.add(str(raw_id).strip())
            continue
        if identifier not in all_valid_ids:
            all_valid_ids.add(identifier)
            all_valid_order.append(identifier)
        if depth_filter:
            raw_depth = original.get(depth_filter.column, "").strip()
            if not raw_depth:
                if not depth_filter.include_blank:
                    item.load_status = "filtered_blank_depth"
                    item.detail = (
                        f"Blank {depth_filter.column} excluded by the depth filter."
                    )
                    depth_filtered_rows += 1
                    continue
            else:
                try:
                    depth = float(raw_depth.replace(",", ""))
                except ValueError:
                    item.load_status = "filtered_invalid_depth"
                    item.detail = (
                        f"{depth_filter.column} is not a numeric depth: {raw_depth!r}."
                    )
                    invalid_depth_rows += 1
                    continue
                keep = (
                    depth < depth_filter.threshold
                    if depth_filter.comparison == "less_than"
                    else depth > depth_filter.threshold
                )
                if not keep:
                    symbol = "<" if depth_filter.comparison == "less_than" else ">"
                    item.load_status = "filtered_depth"
                    item.detail = (
                        f"Depth {depth:g} ft does not satisfy {symbol} "
                        f"{depth_filter.threshold:g} ft."
                    )
                    depth_filtered_rows += 1
                    continue
        eligible_ids.add(identifier)
        if identifier not in variants_by_id:
            valid_order.append(identifier)
            variants_by_id[identifier] = variants
        rows_by_id[identifier].append(len(audit) - 1)

    candidates: OrderedDict[tuple[str, str], _LogicalCandidate] = OrderedDict()
    placeholder_indices: list[int] = []
    duplicate_references = 0

    for identifier in valid_order:
        variants = variants_by_id[identifier]
        for folder_path in folder_paths:
            if _name_matches_identifier(folder_path, variants):
                basename = os.path.basename(folder_path).casefold()
                key = (identifier, basename)
                candidates.setdefault(key, _LogicalCandidate(identifier, basename))
                candidates[key].folder_path = folder_path

    seen_row_refs: set[tuple[str, str, str]] = set()
    source_dir = os.path.dirname(os.path.abspath(table.source_path))

    def listed_path(value: str) -> str:
        return value if os.path.isabs(value) else os.path.join(source_dir, value)

    for identifier, indices in rows_by_id.items():
        for index in indices:
            item = audit[index]
            raw_path = item.source_tif_path.strip()
            if raw_path.casefold() in PLACEHOLDERS:
                placeholder_indices.append(index)
                continue
            if not raw_path:
                continue
            basename = os.path.basename(raw_path.replace("/", os.sep)).casefold()
            if not basename:
                continue
            key = (identifier, basename)
            candidate = candidates.setdefault(
                key, _LogicalCandidate(identifier, basename)
            )
            candidate.row_indices.append(index)
            resolved_listing = listed_path(raw_path)
            ref_key = (identifier, basename,
                       os.path.normcase(os.path.abspath(resolved_listing)))
            if ref_key in seen_row_refs:
                duplicate_references += 1
            else:
                seen_row_refs.add(ref_key)
                if candidate.listed_paths:
                    duplicate_references += 1
                candidate.listed_paths.append(resolved_listing)
            if basename in folder_by_name:
                candidate.folder_path = folder_by_name[basename]

    basename_owner: dict[str, str] = {}
    conflict_keys: set[tuple[str, str]] = set()
    for key, candidate in candidates.items():
        owner = basename_owner.setdefault(candidate.basename, candidate.identifier)
        if owner != candidate.identifier:
            conflict_keys.add(key)

    failed_by_id: dict[str, int] = defaultdict(int)
    available_roots: dict[str, tuple[bool, str]] = {}
    for candidate in candidates.values():
        candidate_paths = [candidate.folder_path] if candidate.folder_path else []
        candidate_paths.extend(candidate.listed_paths)
        for path in candidate_paths:
            root = _unc_share_root(path)
            if root:
                available_roots.setdefault(root, (False, ""))
    total_candidates = sum(key not in conflict_keys for key in candidates)
    for root in list(available_roots):
        if cancel_event and cancel_event.is_set():
            raise ImportCancelled("Image-list loading was cancelled.")
        if progress_fn:
            progress_fn(0, total_candidates, f"Checking network share {root}")
        available_roots[root] = _network_root_status(
            root, network_timeout, network_check
        )

    unavailable_paths = {
        os.path.normcase(os.path.abspath(path))
        for candidate in candidates.values()
        for path in ([candidate.folder_path] if candidate.folder_path else [])
                    + candidate.listed_paths
        if (_unc_share_root(path)
            and not available_roots[_unc_share_root(path)][0])
    }

    def resolve_candidate(candidate: _LogicalCandidate) -> None:
        if cancel_event and cancel_event.is_set():
            return
        choices: list[tuple[str, str]] = []
        if candidate.folder_path:
            choices.append((candidate.folder_path, "selected_folder"))
        for path in candidate.listed_paths:
            if not any(os.path.normcase(os.path.abspath(path)) ==
                       os.path.normcase(os.path.abspath(existing))
                       for existing, _ in choices):
                choices.append((path, "listed_path"))
        failures: list[tuple[str, str]] = []
        for path, source in choices:
            if cancel_event and cancel_event.is_set():
                return
            root = _unc_share_root(path)
            if root and not available_roots[root][0]:
                failures.append(("network_unavailable", available_roots[root][1]))
                continue
            ok, status, detail = _probe(path, probe_fn)
            if ok:
                candidate.resolved_path = os.path.abspath(path)
                candidate.resolution_source = source
                return
            failures.append((status, detail))
        if failures:
            candidate.failure_status, candidate.failure_detail = failures[-1]
        elif candidate.listed_paths:
            candidate.failure_status = "path_missing"
            candidate.failure_detail = "No listed path was accessible."
        else:
            candidate.failure_status = "no_path"
            candidate.failure_detail = "No TIFF path was supplied or matched."

    candidate_items = [
        (key, candidate) for key, candidate in candidates.items()
        if key not in conflict_keys
    ]
    executor = ThreadPoolExecutor(
        max_workers=max(1, min(int(max_workers), len(candidate_items) or 1)),
        thread_name_prefix="logreview-manifest",
    )
    futures = {
        executor.submit(resolve_candidate, candidate): candidate
        for _key, candidate in candidate_items
    }
    completed = 0
    cancelled = False
    try:
        for future in as_completed(futures):
            future.result()
            completed += 1
            candidate = futures[future]
            if progress_fn:
                progress_fn(completed, len(candidate_items), candidate.basename)
            if cancel_event and cancel_event.is_set():
                cancelled = True
                break
    finally:
        executor.shutdown(wait=not cancelled, cancel_futures=cancelled)
    if cancelled:
        raise ImportCancelled("Image-list loading was cancelled.")

    resolved_candidates: list[_LogicalCandidate] = []
    for key, candidate in candidates.items():
        if key in conflict_keys:
            candidate.failure_status = "basename_conflict"
            candidate.failure_detail = (
                f"The same TIFF basename is already assigned to identifier "
                f"{basename_owner[candidate.basename]}."
            )
            failed_by_id[candidate.identifier] += 1
            continue
        if candidate.resolved_path:
            resolved_candidates.append(candidate)
        else:
            failed_by_id[candidate.identifier] += 1

    resolved_by_id: dict[str, list[_LogicalCandidate]] = defaultdict(list)
    candidate_by_key = {(c.identifier, c.basename): c for c in candidates.values()}
    for candidate in resolved_candidates:
        resolved_by_id[candidate.identifier].append(candidate)

    for identifier, indices in rows_by_id.items():
        loaded = bool(resolved_by_id.get(identifier))
        for index in indices:
            item = audit[index]
            item.well_loaded = loaded
            raw_path = item.source_tif_path.strip()
            if raw_path.casefold() in PLACEHOLDERS:
                item.load_status = "placeholder_covered" if loaded else "placeholder_unresolved"
                item.detail = (
                    "Another TIFF loaded for this identifier."
                    if loaded else "No TIFF was available for this identifier."
                )
                continue
            if not raw_path:
                if loaded:
                    resolved = resolved_by_id[identifier]
                    item.resolved_tif_path = "; ".join(
                        candidate.resolved_path for candidate in resolved
                    )
                    item.row_image_loaded = True
                    item.resolution_source = "selected_folder"
                    item.load_status = "loaded_from_selected_folder"
                    item.detail = (
                        f"Matched {len(resolved)} image(s) by identifier in the "
                        "selected image collection."
                    )
                else:
                    item.load_status = "no_path_unresolved"
                    item.detail = "No TIFF path was supplied or matched."
                continue
            basename = os.path.basename(raw_path.replace("/", os.sep)).casefold()
            candidate = candidate_by_key.get((identifier, basename))
            if candidate and candidate.resolved_path:
                item.resolved_tif_path = candidate.resolved_path
                item.row_image_loaded = True
                item.resolution_source = candidate.resolution_source
                if candidate.resolution_source == "selected_folder":
                    item.load_status = "loaded_from_selected_folder"
                elif os.path.normcase(os.path.abspath(listed_path(raw_path))) == os.path.normcase(
                        candidate.resolved_path):
                    item.load_status = "recovered_from_listed_path"
                else:
                    item.load_status = "duplicate_reference_loaded"
                continue
            status = candidate.failure_status if candidate else "path_missing"
            detail = candidate.failure_detail if candidate else "No TIFF candidate was resolved."
            item.load_status = status + ("_well_covered" if loaded else "_unresolved")
            item.detail = detail

    loaded_ids = [identifier for identifier in valid_order if resolved_by_id.get(identifier)]
    ordered_paths: list[str] = []
    path_identifiers: dict[str, str] = {}
    metadata_columns = tuple((name, column) for name, column in (
        ("latitude", table.latitude_column),
        ("longitude", table.longitude_column),
        ("log_top_depth", table.log_top_depth_column),
        ("log_bottom_depth", table.log_bottom_depth_column),
    ) if column)
    identifier_locations: dict[str, dict[str, str]] = {}
    for identifier, indices in rows_by_id.items():
        location = {name: "" for name, _column in metadata_columns}
        for index in indices:
            original = audit[index].original
            for name, column in metadata_columns:
                if column and not location[name]:
                    location[name] = original.get(column, "").strip()
        identifier_locations[identifier] = location
    path_locations: dict[str, dict[str, str]] = {}
    for identifier in valid_order:
        for candidate in resolved_by_id.get(identifier, []):
            ordered_paths.append(candidate.resolved_path)
            path_key = os.path.normcase(os.path.abspath(candidate.resolved_path))
            path_identifiers[path_key] = identifier
            fallback = identifier_locations.get(identifier, {})
            specific = {name: "" for name, _column in metadata_columns}
            # When source rows identify a particular TIFF, their interval values
            # take precedence over the identifier-level fallback. This preserves
            # different available intervals for multiple logs from one well.
            for index in candidate.row_indices:
                original = audit[index].original
                for name, column in metadata_columns:
                    value = original.get(column, "").strip() if column else ""
                    if value and not specific[name]:
                        specific[name] = value
            location = {
                name: specific[name] or fallback.get(name, "")
                for name, _column in metadata_columns
            }
            path_locations[path_key] = location

    # A filtered row can refer to an identifier retained by another source row.
    # Keep its audit status as filtered while still reporting whether the well loaded.
    loaded_id_set = set(loaded_ids)
    for item in audit:
        identifier, _ = normalize_identifier(item.source_identifier)
        if item.load_status.startswith("filtered_"):
            item.well_loaded = identifier in loaded_id_set

    placeholder_covered = sum(
        1 for index in placeholder_indices if audit[index].load_status == "placeholder_covered"
    )
    stats = ImportStats(
        input_rows=len(audit),
        valid_unique_identifiers=len(all_valid_order),
        invalid_unique_identifiers=len(invalid_values),
        unique_tiff_candidates=len(candidates),
        loaded_tiffs=len(resolved_candidates),
        loaded_from_folder=sum(c.resolution_source == "selected_folder"
                               for c in resolved_candidates),
        recovered_from_paths=sum(c.resolution_source == "listed_path"
                                 for c in resolved_candidates),
        duplicate_references=duplicate_references,
        loaded_identifiers=len(loaded_ids),
        partially_loaded_identifiers=sum(
            bool(resolved_by_id.get(identifier)) and failed_by_id[identifier] > 0
            for identifier in valid_order
        ),
        unrecovered_identifiers=sum(not resolved_by_id.get(identifier)
                                    for identifier in valid_order),
        placeholder_rows=len(placeholder_indices),
        placeholder_covered=placeholder_covered,
        placeholder_unresolved=len(placeholder_indices) - placeholder_covered,
        eligible_unique_identifiers=len(eligible_ids),
        depth_filtered_rows=depth_filtered_rows,
        depth_filtered_identifiers=len(all_valid_ids - eligible_ids),
        invalid_depth_rows=invalid_depth_rows,
        depth_filter_description=depth_filter.description if depth_filter else "",
        unavailable_network_shares=sum(not available for available, _ in
                                       available_roots.values()),
        network_paths_skipped=len(unavailable_paths),
    )
    return ImportResult(
        table, ordered_paths, path_identifiers, path_locations, loaded_ids, audit, stats
    )


def audit_report_path(source_path: str, workbook_path: str,
                      now: _dt.datetime | None = None) -> str:
    stamp = (now or _dt.datetime.now()).strftime("%Y%m%d_%H%M%S")
    stem = Path(source_path).stem or "image_list"
    return os.path.join(os.path.dirname(os.path.abspath(workbook_path)),
                        f"{stem}_load_status_{stamp}.csv")


def write_audit_csv(path: str, result: ImportResult) -> str:
    """Write a source-preserving row-level load report."""
    folder = os.path.dirname(os.path.abspath(path))
    if folder:
        os.makedirs(folder, exist_ok=True)
    headers = list(result.source.headers) + AUDIT_COLUMNS
    with open(path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in result.audit_rows:
            writer.writerow(row.as_dict())
    return os.path.abspath(path)
