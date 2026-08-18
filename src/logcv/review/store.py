"""The review workbook: one row per log, resumable, plus the stamp-type list.

Design rules that matter downstream:

* The workbook is the **only** state. Re-opening the app reads it back and the
  review carries on; nothing is kept in a side file.
* Every save rewrites the sheet from scratch, so the file on disk is always the
  whole current answer -- no partial appends to reconcile.
* `file_link` is a clickable hyperlink for a human, and `file_path` repeats the
  same path as plain text because a hyperlink is invisible to pandas.
* Stamp types live in an editable JSON next to the workbook. Adding "TWDB" or
  "Railroad Commission" is one line there, or one dialog in the app -- no code.
"""
from __future__ import annotations

import datetime as _dt
import importlib.resources
import json
import os
import tempfile
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

#: Sheet the review rows live on.
SHEET = "review"
#: Optional sheet mirroring the selected review rows for subset mode.
SUBSET_SHEET = "subset"
#: Review header row, in order.
COLUMNS = [
    "log_api",        # 14-digit API parsed from the file name
    "file_link",      # hyperlink to the image, shows the file name
    "has_stamp",      # TRUE / FALSE / blank when not yet reviewed
    "type_of_stamp",  # comma-separated stamp types; blank when FALSE
    "log_types",      # comma-separated fixed log classifications
    "notes",          # free text from the reviewer
    "reviewed_at",    # ISO timestamp of the latest committed edit
    "reviewed_by",    # reviewer responsible for the latest committed edit
    "file_path",      # same target as file_link, readable by pandas
]
SUBSET_COLUMNS = list(COLUMNS)
LEGACY_SUBSET_COLUMNS = ["position", "log_api"]
#: Required in all supported schemas. ``api14`` is accepted as the legacy alias
#: for ``log_api``; ``log_types`` and ``reviewed_by`` were added later.
REQUIRED_COLUMNS = [
    "file_link", "file_path", "has_stamp", "type_of_stamp", "notes", "reviewed_at"
]
#: What the app offers before anyone adds anything.
DEFAULT_STAMP_TYPES = ["IHS"]
#: Name of the editable list, kept beside the workbook.
STAMP_TYPES_FILE = "stamp_types.json"
DEFAULT_LOG_TYPES = [
    "Caliper",
    "Gamma Ray",
    "Porosity",
    "Resistivity deep",
    "Resistivity medium",
    "Resistivity shallow",
    "Sonic",
    "Spontaneous Potential",
]

_TRUE = {"true", "yes", "y", "1", "t"}
_FALSE = {"false", "no", "n", "0", "f"}


class WorkbookLocked(OSError):
    """The workbook is open in Excel, so it cannot be written."""


class InvalidWorkbook(ValueError):
    """A selected Excel file is not a compatible LogReview workbook."""


@dataclass(frozen=True)
class WorkbookInspection:
    """Structurally valid workbook/folder filename reconciliation."""

    workbook_names: tuple[str, ...]
    missing_names: tuple[str, ...]
    extra_names: tuple[str, ...]


def split_values(value: str | None) -> list[str]:
    """Split a comma-separated workbook cell, trimmed and de-duplicated."""
    out: list[str] = []
    seen: set[str] = set()
    for item in str(value or "").split(","):
        clean = item.strip()
        key = clean.casefold()
        if clean and key not in seen:
            out.append(clean)
            seen.add(key)
    return out


def join_values(values) -> str:
    """Canonical comma-separated representation used in Excel."""
    return ", ".join(split_values(",".join(str(value) for value in values)))


def load_log_types() -> list[str]:
    """Load the fixed choices bundled with the application."""
    try:
        resource = importlib.resources.files(__package__).joinpath("log_types.json")
        data = json.loads(resource.read_text(encoding="utf-8"))
        values = data.get("types") if isinstance(data, dict) else data
        if isinstance(values, list):
            cleaned = split_values(",".join(str(value) for value in values))
            if cleaned:
                return cleaned
    except (OSError, ValueError, TypeError):
        pass
    return list(DEFAULT_LOG_TYPES)


@dataclass
class Record:
    """One log's verdict."""

    file_name: str
    file_path: str
    has_stamp: bool | None = None
    stamp_type: str = ""
    log_types: str = ""
    notes: str = ""
    api14: str = ""
    reviewed_at: str = ""
    reviewed_by: str = ""

    @property
    def reviewed(self) -> bool:
        """The independent stamp-review stage is complete."""
        return self.has_stamp is False or (
            self.has_stamp is True and bool(split_values(self.stamp_type))
        )

    @property
    def has_entry(self) -> bool:
        """Some reviewer-supplied information has been committed."""
        return self.has_stamp is not None or bool(
            self.stamp_type or self.log_types or self.notes
        )

    @property
    def incomplete(self) -> bool:
        """A committed entry exists, but the stamp stage is unfinished."""
        return self.has_entry and not self.reviewed

    def set_verdict(self, has_stamp: bool | None, stamp_type: str = "") -> None:
        self.has_stamp = has_stamp
        self.stamp_type = join_values(split_values(stamp_type)) if has_stamp else ""
        self.reviewed_at = (
            "" if has_stamp is None
            else _dt.datetime.now().replace(microsecond=0).isoformat(sep=" ")
        )


def _parse_bool(value) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in _TRUE:
        return True
    if text in _FALSE:
        return False
    return None


def _read_subset_apis(sheet) -> list[str]:
    """Read subset membership from the full mirror or v1.6 legacy schema."""
    rows = sheet.iter_rows(values_only=True)
    try:
        first = next(rows)
    except StopIteration as exc:
        raise InvalidWorkbook("The subset worksheet is empty.") from exc
    header = {
        str(value).strip().lower(): i
        for i, value in enumerate(first) if value is not None
    }
    if all(name in header for name in LEGACY_SUBSET_COLUMNS):
        return _read_legacy_subset_apis(rows, header)

    actual = [str(value).strip().lower() for value in first if value is not None]
    if actual != SUBSET_COLUMNS:
        raise InvalidWorkbook(
            "The subset worksheet must use the same columns and order as the review "
            "worksheet."
        )
    apis: list[str] = []
    seen_apis: set[str] = set()
    seen_names: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        api = str(row[header["log_api"]] or "").strip()
        name = str(row[header["file_link"]] or "").strip()
        if not api and not name and not any(value not in (None, "") for value in row):
            continue
        if len(api) != 14 or not api.isdigit():
            raise InvalidWorkbook(
                f"Subset row {row_number} must contain a 14-digit log_api."
            )
        if not name:
            raise InvalidWorkbook(f"Subset row {row_number} has no file_link value.")
        name_key = name.casefold()
        if name_key in seen_names:
            raise InvalidWorkbook(
                f"Subset row {row_number} duplicates file name {name!r}."
            )
        seen_names.add(name_key)
        if api not in seen_apis:
            apis.append(api)
            seen_apis.add(api)
    return apis


def _read_legacy_subset_apis(rows, header: dict[str, int]) -> list[str]:
    """Read the v1.6/v1.6.1 position/log_api membership-only sheet."""
    positioned: list[tuple[int, str]] = []
    seen_positions: set[int] = set()
    seen_apis: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        def value(column: str):
            index = header[column]
            return row[index] if index < len(row) else None

        raw_position = value("position")
        raw_api = value("log_api")
        if raw_position in (None, "") and raw_api in (None, ""):
            continue
        try:
            position = int(raw_position)
        except (TypeError, ValueError) as exc:
            raise InvalidWorkbook(
                f"Subset row {row_number} has an invalid position."
            ) from exc
        api = str(raw_api or "").strip()
        if position < 1 or position in seen_positions:
            raise InvalidWorkbook(
                f"Subset row {row_number} has a duplicate or invalid position."
            )
        if len(api) != 14 or not api.isdigit() or api in seen_apis:
            raise InvalidWorkbook(
                f"Subset row {row_number} must contain a unique 14-digit log_api."
            )
        seen_positions.add(position)
        seen_apis.add(api)
        positioned.append((position, api))
    return [api for _, api in sorted(positioned)]


def validate_review_workbook(
    path: str, image_names: set[str] | None = None
) -> WorkbookInspection:
    """Raise :class:`InvalidWorkbook` unless ``path`` is safe to resume.

    Validation is intentionally stricter than ``ReviewStore.load``: a user is
    selecting an arbitrary Excel file through a dialog, so silently treating a
    random first worksheet as review state would be dangerous.
    """
    issues: list[str] = []
    try:
        book = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise InvalidWorkbook(f"Excel could not read this file: {exc}") from exc
    try:
        if SHEET not in book.sheetnames:
            raise InvalidWorkbook(f"Missing required worksheet {SHEET!r}.")
        sheet = book[SHEET]
        rows = sheet.iter_rows(values_only=True)
        try:
            first = next(rows)
        except StopIteration:
            raise InvalidWorkbook("The review worksheet is empty.")
        header = {
            str(value).strip().lower(): i
            for i, value in enumerate(first) if value is not None
        }
        missing = [name for name in REQUIRED_COLUMNS if name not in header]
        if "log_api" not in header and "api14" not in header:
            missing.append("log_api (or legacy api14)")
        if missing:
            issues.append("Missing columns: " + ", ".join(missing))

        # Row-level checks require the three core columns. Missing headers are
        # already enough to reject the file, so avoid secondary index errors.
        core = {"file_link", "has_stamp", "type_of_stamp"}
        if core.issubset(header):
            seen: set[str] = set()
            workbook_names: list[str] = []
            for row_number, row in enumerate(rows, start=2):
                def value(column: str):
                    index = header[column]
                    return row[index] if index < len(row) else None

                name = str(value("file_link") or "").strip()
                if not name:
                    if any(cell not in (None, "") for cell in row):
                        issues.append(f"Row {row_number} has data but no file name.")
                    continue
                key = name.casefold()
                if key in seen:
                    issues.append(f"Duplicate file name at row {row_number}: {name}")
                seen.add(key)
                workbook_names.append(name)
                raw_verdict = value("has_stamp")
                if raw_verdict not in (None, "") and _parse_bool(raw_verdict) is None:
                    issues.append(
                        f"Row {row_number} has an invalid has_stamp value: {raw_verdict!r}"
                    )
                if _parse_bool(raw_verdict) is False and str(
                    value("type_of_stamp") or ""
                ).strip():
                    issues.append(
                        f"Row {row_number} says no stamp but also supplies a stamp type."
                    )
        else:
            workbook_names = []
        if SUBSET_SHEET in book.sheetnames:
            _read_subset_apis(book[SUBSET_SHEET])
    finally:
        book.close()

    if issues:
        raise InvalidWorkbook("\n".join(issues[:12]))

    if image_names is None:
        return WorkbookInspection(tuple(workbook_names), (), ())
    folder_by_key = {name.casefold(): name for name in image_names}
    workbook_by_key = {name.casefold(): name for name in workbook_names}
    missing_names = tuple(
        folder_by_key[key] for key in sorted(folder_by_key.keys() - workbook_by_key.keys())
    )
    extra_names = tuple(
        workbook_by_key[key] for key in sorted(workbook_by_key.keys() - folder_by_key.keys())
    )
    return WorkbookInspection(tuple(workbook_names), missing_names, extra_names)


class ReviewStore:
    """Records keyed by file name, loaded from and saved to one workbook."""

    def __init__(self, path: str):
        self.path = os.path.abspath(path)
        self.records: dict[str, Record] = {}
        self.subset_apis: list[str] = []
        self.dirty = False

    # ------------------------------------------------------------------ load

    def load(self) -> int:
        """Read existing verdicts. Returns how many rows were found."""
        self.records.clear()
        self.subset_apis.clear()
        if not os.path.exists(self.path):
            return 0
        book = load_workbook(self.path, data_only=True)
        sheet = book[SHEET] if SHEET in book.sheetnames else book.worksheets[0]
        rows = sheet.iter_rows(values_only=False)
        try:
            header_cells = next(rows)
        except StopIteration:
            book.close()
            return 0
        header = {
            str(cell.value).strip().lower(): i
            for i, cell in enumerate(header_cells) if cell.value
        }

        def cell(row, key, *aliases):
            i = next((header.get(name) for name in (key, *aliases)
                      if header.get(name) is not None), None)
            return row[i] if i is not None and i < len(row) else None

        found = 0
        for row in rows:
            link_cell = cell(row, "file_link")
            path_cell = cell(row, "file_path")
            path = str(path_cell.value).strip() if path_cell and path_cell.value else ""
            name = ""
            if link_cell is not None and link_cell.value:
                name = str(link_cell.value).strip()
            if not name and path:
                name = os.path.basename(path)
            if not name:
                continue
            if not path and link_cell is not None and link_cell.hyperlink:
                path = link_cell.hyperlink.target or ""
            record = Record(
                file_name=name,
                file_path=path,
                has_stamp=_parse_bool(getattr(cell(row, "has_stamp"), "value", None)),
                stamp_type=str(getattr(cell(row, "type_of_stamp"), "value", "") or "").strip(),
                log_types=str(getattr(cell(row, "log_types"), "value", "") or "").strip(),
                notes=str(getattr(cell(row, "notes"), "value", "") or "").strip(),
                api14=str(getattr(cell(row, "log_api", "api14"), "value", "") or "").strip(),
                reviewed_at=str(getattr(cell(row, "reviewed_at"), "value", "") or "").strip(),
                reviewed_by=str(getattr(cell(row, "reviewed_by"), "value", "") or "").strip(),
            )
            self.records[name] = record
            found += 1
        if SUBSET_SHEET in book.sheetnames:
            self.subset_apis = _read_subset_apis(book[SUBSET_SHEET])
        book.close()
        return found

    # ----------------------------------------------------------------- rows

    def record_for(self, path: str, api14: str = "") -> Record:
        """The record for ``path``, created blank the first time it is asked for."""
        name = os.path.basename(path)
        record = self.records.get(name)
        if record is None:
            matched = next((key for key in self.records if key.casefold() == name.casefold()), None)
            if matched is not None:
                record = self.records.pop(matched)
                record.file_name = name
                self.records[name] = record
        if record is None:
            record = Record(file_name=name, file_path=os.path.abspath(path), api14=api14)
            self.records[name] = record
        else:
            # A folder move rewrites the path; the verdict still belongs to the name.
            record.file_path = os.path.abspath(path)
            if api14 and not record.api14:
                record.api14 = api14
        return record

    def stamp_types_in_use(self) -> list[str]:
        found: dict[str, str] = {}
        for record in self.records.values():
            for value in split_values(record.stamp_type):
                found.setdefault(value.casefold(), value)
        return sorted(found.values(), key=str.casefold)

    def counts(self, names: list[str] | None = None) -> tuple[int, int, int]:
        """(reviewed, with a stamp, incomplete)."""
        if names is None:
            values = list(self.records.values())
        else:
            wanted = {name.casefold() for name in names}
            values = [record for name, record in self.records.items()
                      if name.casefold() in wanted]
        return (
            sum(1 for r in values if r.reviewed),
            sum(1 for r in values if r.has_stamp is True),
            sum(1 for r in values if r.incomplete),
        )

    def most_recent_reviewer(self) -> str:
        """Reviewer on the lexically latest ISO timestamp, if any."""
        candidates = [record for record in self.records.values() if record.reviewed_by]
        if not candidates:
            return ""
        return max(candidates, key=lambda record: record.reviewed_at or "").reviewed_by

    def remove_records(self, names: list[str] | tuple[str, ...]) -> dict[str, Record]:
        """Remove case-insensitively named rows and return them for rollback."""
        wanted = {name.casefold() for name in names}
        removed = {name: record for name, record in self.records.items()
                   if name.casefold() in wanted}
        for name in removed:
            self.records.pop(name)
        return removed

    def replace_subset(self, api_numbers: list[str]) -> None:
        """Replace ordered subset membership, de-duplicating APIs in place."""
        ordered: list[str] = []
        seen: set[str] = set()
        for value in api_numbers:
            api = str(value).strip()
            if len(api) != 14 or not api.isdigit():
                raise ValueError(f"Subset API must contain exactly 14 digits: {value!r}")
            if api not in seen:
                ordered.append(api)
                seen.add(api)
        self.subset_apis = ordered
        self.dirty = True

    # ------------------------------------------------------------------ save

    def save(self, order: list[str] | None = None) -> str:
        """Rewrite the workbook. ``order`` is a list of file names for row order."""
        if order:
            requested = {name.casefold() for name in order}
            extras = [name for name in self.records if name.casefold() not in requested]
            names = list(order) + extras
        else:
            names = sorted(self.records)
        book = Workbook()
        sheet = book.active
        sheet.title = SHEET

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="44546A")
        link_font = Font(color="0563C1", underline="single")
        widths = [16, 46, 11, 24, 28, 40, 20, 16, 70]

        def write_record_sheet(target, record_names: list[str]) -> None:
            for col, column_name in enumerate(COLUMNS, start=1):
                cell = target.cell(row=1, column=col, value=column_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            written = 0
            for name in record_names:
                record = self.records.get(name)
                if record is None:
                    continue
                written += 1
                row_i = written + 1
                api_cell = target.cell(row=row_i, column=1, value=record.api14 or None)
                api_cell.number_format = "@"
                link = target.cell(row=row_i, column=2, value=record.file_name)
                if record.file_path:
                    link.hyperlink = record.file_path
                    link.font = link_font
                target.cell(row=row_i, column=3, value=record.has_stamp)
                target.cell(row=row_i, column=4, value=record.stamp_type or None)
                target.cell(row=row_i, column=5, value=record.log_types or None)
                target.cell(row=row_i, column=6, value=record.notes or None)
                target.cell(row=row_i, column=7, value=record.reviewed_at or None)
                target.cell(row=row_i, column=8, value=record.reviewed_by or None)
                target.cell(row=row_i, column=9, value=record.file_path or None)
            for col, width in enumerate(widths, start=1):
                target.column_dimensions[get_column_letter(col)].width = width
            target.freeze_panes = "A2"
            last = max(2, written + 1)
            target.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last}"

        write_record_sheet(sheet, names)

        if self.subset_apis:
            subset = book.create_sheet(SUBSET_SHEET)
            subset_source = list(order) if order else names
            wanted = set(self.subset_apis)
            subset_names = [
                name for name in subset_source
                if (self.records.get(name) is not None
                    and self.records[name].api14 in wanted)
            ]
            write_record_sheet(subset, subset_names)

        folder = os.path.dirname(self.path)
        if folder:
            os.makedirs(folder, exist_ok=True)
        temp_path = ""
        try:
            descriptor, temp_path = tempfile.mkstemp(
                prefix=".logreview_", suffix=".xlsx", dir=folder or "."
            )
            os.close(descriptor)
            book.save(temp_path)
            os.replace(temp_path, self.path)
            temp_path = ""
        except PermissionError as exc:
            raise WorkbookLocked(
                f"cannot write {self.path} -- it is open in Excel. Close it and save again."
            ) from exc
        finally:
            book.close()
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except OSError:
                    pass
        self.dirty = False
        return self.path


# ------------------------------------------------------------------ stamp types


def stamp_types_path(workbook: str) -> str:
    return os.path.join(os.path.dirname(os.path.abspath(workbook)), STAMP_TYPES_FILE)


def load_stamp_types(path: str) -> list[str]:
    """The stamp types on offer. Missing or broken file -> the built-in default."""
    try:
        with open(path, encoding="utf-8") as handle:
            data = json.load(handle)
    except (OSError, ValueError):
        return list(DEFAULT_STAMP_TYPES)
    types = data.get("types") if isinstance(data, dict) else data
    if not isinstance(types, list):
        return list(DEFAULT_STAMP_TYPES)
    cleaned = [str(t).strip() for t in types if str(t).strip()]
    return cleaned or list(DEFAULT_STAMP_TYPES)


def save_stamp_types(path: str, types: list[str]) -> None:
    """Persist the list, de-duplicated, order preserved."""
    seen, ordered = set(), []
    for t in types:
        t = str(t).strip()
        if t and t.lower() not in seen:
            seen.add(t.lower())
            ordered.append(t)
    folder = os.path.dirname(os.path.abspath(path))
    if folder:
        os.makedirs(folder, exist_ok=True)
    with open(path, "w", encoding="utf-8") as handle:
        json.dump({"types": ordered}, handle, indent=2)
        handle.write("\n")
