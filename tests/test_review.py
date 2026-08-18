"""Unit tests for the hand-review app's headless storage and rendering layers.

The Tk window is not tested here (it needs a display); what is tested is
everything a wrong answer would silently corrupt:

* the workbook round-trip, including resume -- if `has_stamp` came back as the
  string "FALSE" instead of `False`, every reviewed "no" would read as a "yes";
* boolean coercion of the values Excel can hand back (bool, 0/1, text);
* viewport geometry, where an off-by-a-factor gives a picture that looks fine but
  is not the rectangle the user is panning to.

Run:  pytest tests -q     (from the project root, with src/ on PYTHONPATH)
"""
from __future__ import annotations

import os
import sys

import numpy as np
import pytest
from PIL import Image

from logcv.review import pages
from logcv.review.pyramid import TilePyramid
from logcv.review.store import (
    COLUMNS,
    DEFAULT_LOG_TYPES,
    DEFAULT_STAMP_TYPES,
    InvalidWorkbook,
    Record,
    ReviewStore,
    SUBSET_COLUMNS,
    SUBSET_SHEET,
    WorkbookLocked,
    join_values,
    load_log_types,
    load_stamp_types,
    save_stamp_types,
    split_values,
    stamp_types_path,
    validate_review_workbook,
)
from logcv.units import Box


# ------------------------------------------------------------------ the record


def test_record_starts_unreviewed():
    record = Record(file_name="a.tif", file_path="C:/a.tif")
    assert record.has_stamp is None
    assert not record.reviewed
    assert not record.incomplete


def test_yes_without_a_type_is_incomplete_until_a_type_is_added():
    record = Record(file_name="a.tif", file_path="C:/a.tif")
    record.set_verdict(True)
    assert not record.reviewed and record.incomplete
    record.set_verdict(True, "IHS")
    assert record.reviewed and not record.incomplete
    assert record.reviewed_at  # stamped with a time


def test_saying_no_clears_any_type_already_picked():
    record = Record(file_name="a.tif", file_path="C:/a.tif")
    record.set_verdict(True, "IHS")
    record.set_verdict(False, "IHS")
    assert record.has_stamp is False
    assert record.stamp_type == ""


# --------------------------------------------------------------- the workbook


def test_workbook_round_trip_preserves_the_verdicts(tmp_path):
    book = tmp_path / "stamp_review.xlsx"
    store = ReviewStore(str(book))
    yes = store.record_for(str(tmp_path / "42175000720000_1.TIF"), "42175000720000")
    yes.set_verdict(True, "IHS")
    yes.stamp_type = "IHS, TWDB"
    yes.log_types = "Gamma Ray, Spontaneous Potential"
    yes.notes = "two stamps, one at the tail"
    yes.reviewed_by = "KP"
    no = store.record_for(str(tmp_path / "42175001200000_2.TIF"), "42175001200000")
    no.set_verdict(False)
    store.record_for(str(tmp_path / "42175010740000_3.TIF"))  # untouched
    store.save(order=[r.file_name for r in store.records.values()])

    reloaded = ReviewStore(str(book))
    assert reloaded.load() == 3
    again = reloaded.records["42175000720000_1.TIF"]
    assert again.has_stamp is True
    assert again.stamp_type == "IHS, TWDB"
    assert again.log_types == "Gamma Ray, Spontaneous Potential"
    assert again.notes == "two stamps, one at the tail"
    assert again.api14 == "42175000720000"
    assert again.reviewed_by == "KP"
    assert reloaded.records["42175001200000_2.TIF"].has_stamp is False
    assert reloaded.records["42175010740000_3.TIF"].has_stamp is None
    assert reloaded.counts() == (2, 1, 0)


def test_saved_sheet_has_the_exact_v16_column_order(tmp_path):
    from openpyxl import load_workbook

    book = tmp_path / "stamp_review.xlsx"
    store = ReviewStore(str(book))
    store.record_for(str(tmp_path / "a.tif")).set_verdict(True, "IHS")
    store.save(order=["a.tif"])

    sheet = load_workbook(book)["review"]
    header = [cell.value for cell in sheet[1]]
    assert header == COLUMNS
    assert header == [
        "log_api", "file_link", "has_stamp", "type_of_stamp", "log_types",
        "notes", "reviewed_at", "reviewed_by", "file_path",
    ]
    assert sheet.cell(row=2, column=2).value == "a.tif"
    assert sheet.cell(row=2, column=2).hyperlink.target.endswith("a.tif")
    assert sheet.cell(row=2, column=3).value is True  # a real Excel boolean
    assert sheet.cell(row=2, column=9).value.endswith("a.tif")


@pytest.mark.parametrize(
    "written,expected",
    [(True, True), (False, False), (1, True), (0, False), ("TRUE", True),
     ("no", False), ("y", True), (None, None), ("", None), ("garbage", None)],
)
def test_has_stamp_is_coerced_from_whatever_excel_returns(tmp_path, written, expected):
    from openpyxl import Workbook

    book = tmp_path / "hand_edited.xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.title = "review"
    sheet.append(COLUMNS)
    sheet.append(["", "a.tif", written, "IHS", "", "", "", "",
                  str(tmp_path / "a.tif")])
    wb.save(book)

    store = ReviewStore(str(book))
    store.load()
    assert store.records["a.tif"].has_stamp is expected


def test_reopening_the_same_folder_keeps_the_verdict_but_updates_the_path(tmp_path):
    store = ReviewStore(str(tmp_path / "book.xlsx"))
    store.record_for("D:/old/a.tif").set_verdict(True, "IHS")
    record = store.record_for(str(tmp_path / "a.tif"))
    assert record.has_stamp is True
    assert record.file_path == str(tmp_path / "a.tif")


def test_comma_separated_values_preserve_order_and_remove_duplicates():
    assert split_values("IHS, TWDB, ihs,  Riley ") == ["IHS", "TWDB", "Riley"]
    assert join_values(["Gamma", "Spontaneous Potential", "gamma"]) == (
        "Gamma, Spontaneous Potential"
    )


def test_note_only_entry_is_present_but_incomplete():
    record = Record(file_name="a.tif", file_path="C:/a.tif", notes="illegible header")
    assert record.has_entry
    assert record.incomplete
    assert not record.reviewed


def test_saved_review_workbook_passes_strict_resume_validation(tmp_path):
    book = tmp_path / "review.xlsx"
    store = ReviewStore(str(book))
    store.record_for(str(tmp_path / "a.tif")).set_verdict(False)
    store.save(order=["a.tif"])

    validate_review_workbook(str(book), {"a.tif"})


def test_random_excel_file_is_rejected_as_a_review_workbook(tmp_path):
    from openpyxl import Workbook

    book = tmp_path / "random.xlsx"
    workbook = Workbook()
    workbook.active.append(["well_id", "depth_ft"])
    workbook.save(book)

    with pytest.raises(InvalidWorkbook, match="Missing required worksheet"):
        validate_review_workbook(str(book), {"a.tif"})


def test_workbook_folder_differences_are_reported_not_rejected(tmp_path):
    book = tmp_path / "review.xlsx"
    store = ReviewStore(str(book))
    store.record_for(str(tmp_path / "other.tif")).set_verdict(True, "IHS")
    store.save(order=["other.tif"])

    inspection = validate_review_workbook(str(book), {"a.tif"})
    assert inspection.missing_names == ("a.tif",)
    assert inspection.extra_names == ("other.tif",)


def test_review_workbook_with_invalid_verdict_is_rejected(tmp_path):
    from openpyxl import Workbook

    book = tmp_path / "bad_verdict.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "review"
    sheet.append(COLUMNS)
    sheet.append(["", "a.tif", "maybe", "", "", "", "", "",
                  str(tmp_path / "a.tif")])
    workbook.save(book)

    with pytest.raises(InvalidWorkbook, match="invalid has_stamp"):
        validate_review_workbook(str(book), {"a.tif"})


def test_v12_workbook_without_log_types_loads_and_upgrades_on_save(tmp_path):
    from openpyxl import Workbook, load_workbook

    legacy_columns = [
        "file_link", "has_stamp", "type_of_stamp", "notes", "api14",
        "file_path", "reviewed_at",
    ]
    book = tmp_path / "v12.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "review"
    sheet.append(legacy_columns)
    sheet.append(["a.tif", True, "IHS", "legacy note", "", "C:/a.tif", "2026-08-18"])
    workbook.save(book)

    validate_review_workbook(str(book), {"a.tif"})
    store = ReviewStore(str(book))
    assert store.load() == 1
    assert store.records["a.tif"].notes == "legacy note"
    assert store.records["a.tif"].log_types == ""
    assert store.records["a.tif"].reviewed_by == ""
    store.save(order=["a.tif"])
    assert [cell.value for cell in load_workbook(book)["review"][1]] == COLUMNS


def test_v13_workbook_migrates_api14_and_preserves_log_types(tmp_path):
    from openpyxl import Workbook

    columns = [
        "file_link", "has_stamp", "type_of_stamp", "log_types", "notes",
        "api14", "file_path", "reviewed_at",
    ]
    book = tmp_path / "v13.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "review"
    sheet.append(columns)
    sheet.append(["a.tif", True, "IHS, TWDB", "Gamma", "note",
                  "42175000720000", "C:/a.tif", "2026-08-18 12:00:00"])
    workbook.save(book)

    validate_review_workbook(str(book), {"a.tif"})
    store = ReviewStore(str(book))
    store.load()
    record = store.records["a.tif"]
    assert record.api14 == "42175000720000"
    assert record.stamp_type == "IHS, TWDB"
    assert record.log_types == "Gamma"


def test_filename_reconciliation_is_case_insensitive(tmp_path):
    book = tmp_path / "review.xlsx"
    store = ReviewStore(str(book))
    store.record_for(str(tmp_path / "A.TIF")).set_verdict(False)
    store.save(order=["A.TIF"])

    inspection = validate_review_workbook(str(book), {"a.tif"})
    assert inspection.missing_names == ()
    assert inspection.extra_names == ()


def test_duplicate_workbook_names_are_rejected_case_insensitively(tmp_path):
    from openpyxl import Workbook

    book = tmp_path / "duplicates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "review"
    sheet.append(COLUMNS)
    sheet.append(["", "a.tif", False, "", "", "", "", "", "C:/a.tif"])
    sheet.append(["", "A.TIF", False, "", "", "", "", "", "C:/A.TIF"])
    workbook.save(book)

    with pytest.raises(InvalidWorkbook, match="Duplicate file name"):
        validate_review_workbook(str(book), {"a.tif"})


def test_save_preserves_workbook_only_rows_after_current_folder_rows(tmp_path):
    book = tmp_path / "review.xlsx"
    store = ReviewStore(str(book))
    store.record_for(str(tmp_path / "current.tif")).set_verdict(False)
    store.record_for(str(tmp_path / "preserved.tif")).set_verdict(True, "IHS")
    store.save(order=["current.tif"])

    reloaded = ReviewStore(str(book))
    assert reloaded.load() == 2
    assert list(reloaded.records) == ["current.tif", "preserved.tif"]
    assert reloaded.counts(["current.tif"]) == (1, 0, 0)


def test_subset_membership_round_trips_in_a_separate_sheet(tmp_path):
    from openpyxl import load_workbook

    book = tmp_path / "review.xlsx"
    store = ReviewStore(str(book))
    store.record_for(str(tmp_path / "42175000720000-a.tif"))
    store.record_for(str(tmp_path / "42175000720000-b.tif"))
    store.record_for(str(tmp_path / "42175001200000-c.tif"))
    store.replace_subset([
        "42175001200000", "42175000720000", "42175001200000", "42175999999999",
    ])
    store.save(order=list(store.records))

    workbook = load_workbook(book)
    assert SUBSET_SHEET in workbook.sheetnames
    subset = workbook[SUBSET_SHEET]
    assert [cell.value for cell in subset[1]] == SUBSET_COLUMNS
    assert [subset.cell(row=i, column=2).value for i in range(2, 5)] == [
        "42175001200000", "42175000720000", "42175999999999",
    ]

    reloaded = ReviewStore(str(book))
    reloaded.load()
    assert reloaded.subset_apis == [
        "42175001200000", "42175000720000", "42175999999999",
    ]


def test_legacy_workbook_without_subset_loads_with_no_membership(tmp_path):
    book = tmp_path / "review.xlsx"
    store = ReviewStore(str(book))
    store.record_for(str(tmp_path / "a.tif"))
    store.save(order=["a.tif"])

    reloaded = ReviewStore(str(book))
    reloaded.load()
    assert reloaded.subset_apis == []


def test_malformed_subset_sheet_is_rejected(tmp_path):
    from openpyxl import load_workbook

    book = tmp_path / "review.xlsx"
    store = ReviewStore(str(book))
    store.record_for(str(tmp_path / "a.tif"))
    store.save(order=["a.tif"])
    workbook = load_workbook(book)
    subset = workbook.create_sheet(SUBSET_SHEET)
    subset.append(["position", "wrong_header"])
    subset.append([1, "42175000720000"])
    workbook.save(book)

    with pytest.raises(InvalidWorkbook, match="subset worksheet is missing"):
        validate_review_workbook(str(book), {"a.tif"})


def test_subset_rejects_invalid_api_values(tmp_path):
    store = ReviewStore(str(tmp_path / "review.xlsx"))
    with pytest.raises(ValueError, match="exactly 14 digits"):
        store.replace_subset(["42175"])


def test_remove_records_supports_reconciliation_rollback(tmp_path):
    store = ReviewStore(str(tmp_path / "review.xlsx"))
    store.record_for("C:/a.tif")
    store.record_for("C:/extra.tif")
    removed = store.remove_records(("EXTRA.TIF",))
    assert list(removed) == ["extra.tif"]
    assert list(store.records) == ["a.tif"]


def test_stamp_types_in_existing_cells_are_split_and_merged_case_insensitively(tmp_path):
    store = ReviewStore(str(tmp_path / "review.xlsx"))
    store.record_for("C:/a.tif").stamp_type = "IHS, TWDB"
    store.record_for("C:/b.tif").stamp_type = "twdb, Riley"
    assert store.stamp_types_in_use() == ["IHS", "Riley", "TWDB"]


def test_most_recent_populated_reviewer_is_used_for_prefill(tmp_path):
    store = ReviewStore(str(tmp_path / "review.xlsx"))
    older = store.record_for("C:/a.tif")
    older.reviewed_at = "2026-08-18 09:00:00"
    older.reviewed_by = "KP"
    newer = store.record_for("C:/b.tif")
    newer.reviewed_at = "2026-08-18 11:00:00"
    newer.reviewed_by = "Jane A. Smith"
    assert store.most_recent_reviewer() == "Jane A. Smith"


def test_failed_atomic_replace_leaves_original_workbook_unchanged(tmp_path, monkeypatch):
    book = tmp_path / "review.xlsx"
    store = ReviewStore(str(book))
    store.record_for(str(tmp_path / "a.tif")).set_verdict(False)
    store.save(order=["a.tif"])
    original = book.read_bytes()

    store.records["a.tif"].notes = "unsaved edit"

    def locked_replace(source, target):
        raise PermissionError("simulated Excel lock")

    monkeypatch.setattr(os, "replace", locked_replace)
    with pytest.raises(WorkbookLocked, match="open in Excel"):
        store.save(order=["a.tif"])
    assert book.read_bytes() == original


# -------------------------------------------------------------- stamp types


def test_stamp_types_default_to_ihs_when_no_file_exists(tmp_path):
    assert load_stamp_types(str(tmp_path / "nothing.json")) == DEFAULT_STAMP_TYPES
    assert DEFAULT_STAMP_TYPES == ["IHS"]


def test_adding_a_type_survives_a_round_trip_and_de_duplicates(tmp_path):
    path = stamp_types_path(str(tmp_path / "book.xlsx"))
    save_stamp_types(path, ["IHS", "TWDB", "ihs", " ", "Riley"])
    assert load_stamp_types(path) == ["IHS", "TWDB", "Riley"]


def test_fixed_log_types_are_loaded_from_the_bundled_json():
    assert load_log_types() == DEFAULT_LOG_TYPES
    assert DEFAULT_LOG_TYPES == [
        "Caliper",
        "Gamma Ray",
        "Porosity",
        "Resistivity deep",
        "Resistivity medium",
        "Resistivity shallow",
        "Sonic",
        "Spontaneous Potential",
    ]


def test_subset_api_matching_requires_a_leading_14_digit_number():
    from logcv.review.app import ReviewApp, _parse_subset_api_lines

    assert ReviewApp._leading_api("42175000720000-log.tif") == "42175000720000"
    assert ReviewApp._leading_api("42175000720000_log.tif") == "42175000720000"
    assert ReviewApp._leading_api("prefix-42175000720000.tif") == ""
    assert ReviewApp._leading_api("421750007200001-log.tif") == ""

    requested, duplicates, invalid = _parse_subset_api_lines([
        "42175000720000", "", " 42175001200000 ", "42175000720000", "API",
    ])
    assert requested == ["42175000720000", "42175001200000"]
    assert duplicates == 1
    assert invalid == [(5, "API")]


# ------------------------------------------------------------------ geometry


def _bilevel_tif(path, width=200, height=500):
    """A white page with a black 20 px band at rows 100-120."""
    array = np.full((height, width), 255, dtype=np.uint8)
    array[100:120, :] = 0
    Image.fromarray(array, mode="L").convert("1").save(path, dpi=(400, 400))
    return str(path)


def test_probe_reads_size_and_dpi_without_decoding(tmp_path):
    info = pages.probe(_bilevel_tif(tmp_path / "42175000720000_x.TIF"))
    assert (info.width, info.height) == (200, 500)
    assert info.dpi == pytest.approx(400)
    assert info.api14 == "42175000720000"


def test_api14_is_read_from_the_filename():
    assert pages.api14_from_name("42175000720000_HOU-WL-IMG-1-77489.TIF") == "42175000720000"
    assert pages.api14_from_name("no_number_here.tif") == ""


def test_viewport_is_screen_sized_and_shows_the_page_at_the_asked_for_offset(tmp_path):
    path = _bilevel_tif(tmp_path / "page.tif")
    with pages.open_page(path) as page:
        view = pages.render_viewport(page, vx=0, vy=90, scale=1.0, vw=200, vh=60)
        assert view.size == (200, 60)
        array = np.asarray(view)
        # The ink band starts at page row 100, i.e. screen row 10 at 1:1.
        assert array[0:9].min() == 255
        assert array[12, 100] == 0


def test_zooming_out_keeps_the_ink_visible(tmp_path):
    path = _bilevel_tif(tmp_path / "page.tif")
    with pages.open_page(path) as page:
        view = pages.render_viewport(page, vx=0, vy=0, scale=0.25, vw=50, vh=125)
        assert view.size == (50, 125)
        assert np.asarray(view).min() < 128  # the 20 px band survived 4x decimation


def _dithered_tif(path, width=200, height=200, coverage=0.4, seed=0):
    """A page whose ink is scattered speckle -- what a scanned grey area becomes."""
    rng = np.random.default_rng(seed)
    ink = rng.random((height, width)) < coverage
    array = np.where(ink, 0, 255).astype(np.uint8)
    Image.fromarray(array, mode="L").convert("1").save(path, dpi=(400, 400))
    return str(path)


def test_a_dithered_region_stays_grey_when_zoomed_out(tmp_path):
    """The 42175010740000 defect: 39 % speckle must not render as a black slab."""
    path = _dithered_tif(tmp_path / "dither.tif", coverage=0.4)
    with pages.open_page(path) as page:
        view = pages.render_viewport(page, vx=0, vy=0, scale=0.25, vw=50, vh=50)
        array = np.asarray(view)
    assert (array < 20).mean() < 0.01          # essentially no solid black
    assert 130 < array.mean() < 175            # ~40 % coverage -> mid grey


def test_bold_mode_still_saturates_a_dithered_region(tmp_path):
    """MAX pooling is kept on purpose for hairlines; it is why it is not default."""
    path = _dithered_tif(tmp_path / "dither.tif", coverage=0.4)
    with pages.open_page(path) as page:
        view = pages.render_viewport(page, vx=0, vy=0, scale=0.25, vw=50, vh=50,
                                     mode="max")
    assert (np.asarray(view) < 20).mean() > 0.95


def test_mean_pool_reports_true_coverage():
    ink = np.zeros((4, 4), dtype=bool)
    ink[0, :] = True                                  # a quarter of every cell
    assert pages.mean_pool(ink, 4).tolist() == [[191]]  # 255 * (1 - 0.25)
    assert pages.mean_pool(np.ones((4, 4), bool), 4).tolist() == [[0]]
    assert pages.mean_pool(np.zeros((4, 4), bool), 4).tolist() == [[255]]


def test_mean_pool_does_not_overflow_on_a_big_factor():
    """factor**2 exceeds 16 bits past factor 255; the sums are uint32 for that."""
    ink = np.ones((300, 300), dtype=bool)
    assert pages.mean_pool(ink, 300).tolist() == [[0]]


def test_a_hairline_survives_bold_mode_where_averaging_fades_it(tmp_path):
    array = np.full((200, 200), 255, dtype=np.uint8)
    array[:, 100] = 0  # one-pixel vertical rule
    Image.fromarray(array, mode="L").convert("1").save(tmp_path / "hair.tif",
                                                       dpi=(400, 400))
    with pages.open_page(str(tmp_path / "hair.tif")) as page:
        faint = np.asarray(pages.render_viewport(page, 0, 0, 1 / 16, 12, 12))
        bold = np.asarray(pages.render_viewport(page, 0, 0, 1 / 16, 12, 12, "max"))
    assert faint.min() > 200   # averaged away to near-white
    assert bold.min() == 0     # still black


def test_panning_past_the_edge_pads_white_instead_of_failing(tmp_path):
    path = _bilevel_tif(tmp_path / "page.tif")
    with pages.open_page(path) as page:
        view = pages.render_viewport(page, vx=-100, vy=-50, scale=1.0, vw=200, vh=100)
        assert view.size == (200, 100)
        assert np.asarray(view)[:, 0:100].min() == 255  # left of the sheet


def test_reduced_tiles_are_persistent_and_source_tiff_is_unchanged(tmp_path):
    path = _bilevel_tif(tmp_path / "page.tif")
    cache = tmp_path / "cache"
    before = (os.path.getsize(path), os.stat(path).st_mtime_ns)

    with pages.open_page(path, cache_root=str(cache)) as page:
        first = pages.render_viewport(page, 0, 0, 0.25, 50, 125)

    tiles = list(cache.rglob("*.png"))
    assert tiles
    assert (os.path.getsize(path), os.stat(path).st_mtime_ns) == before

    with pages.open_page(path, cache_root=str(cache)) as page:
        # A second process/session must be able to render from disk alone.
        page._pyramid.reader = lambda *args: (_ for _ in ()).throw(
            AssertionError("source was read despite a valid disk tile")
        )
        second = pages.render_viewport(page, 0, 0, 0.25, 50, 125)
    assert np.array_equal(np.asarray(first), np.asarray(second))


def test_prefetch_populates_rows_above_and_below_the_view(tmp_path):
    source = tmp_path / "source.tif"
    source.write_bytes(b"identity only")
    reads = []

    def reader(box, factor, mode):
        reads.append((box, factor, mode))
        return np.full(((box.height + factor - 1) // factor,
                        (box.width + factor - 1) // factor), 255, np.uint8)

    pyramid = TilePyramid(str(source), 128, 320, None, reader,
                          tile_size=32, memory_limit=32)
    made = pyramid.prefetch_adjacent(Box(0, 96, 64, 160), 1, "mean")

    assert made == 4  # two x tiles in one row above and one row below
    assert {box.y0 for box, _, _ in reads} == {64, 160}


# ------------------------------------------------- packaged-build behaviour


def test_a_packaged_build_asks_for_a_folder_and_saves_beside_itself(tmp_path, monkeypatch):
    """Frozen, there is no project tree: no default corpus, and outputs go next
    to the unpacked exe rather than into `tasks/*/outputs/`."""
    from logcv.review import app as review_app

    exe = tmp_path / "LogReview" / "LogReview.exe"
    monkeypatch.setattr(sys, "frozen", True, raising=False)
    monkeypatch.setattr(sys, "executable", str(exe))

    assert review_app._default_folder() is None
    book = review_app._default_workbook(r"D:\scans\Victoria_TIFs")
    assert book == str(tmp_path / "LogReview" / "reviews"
                       / "Victoria_TIFs_stamp_review.xlsx")
    assert review_app._default_cache_dir() == str(tmp_path / "LogReview" / "cache")


def test_source_build_asks_for_a_folder_and_uses_a_local_reviews_folder(tmp_path, monkeypatch):
    from logcv.review import app as review_app

    monkeypatch.delattr(sys, "frozen", raising=False)
    monkeypatch.chdir(tmp_path)

    assert review_app._default_folder() is None
    book = review_app._default_workbook(r"D:\scans\Victoria_TIFs")
    assert book == str(tmp_path / "reviews" / "Victoria_TIFs_stamp_review.xlsx")
    assert review_app._default_cache_dir() == str(tmp_path / "cache")


def test_factor_for_scale_matches_the_zoom():
    assert pages.factor_for_scale(1.0) == 1
    assert pages.factor_for_scale(2.0) == 1
    assert pages.factor_for_scale(0.5) == 2
    assert pages.factor_for_scale(0.1) == 8
    assert pages.factor_for_scale(1 / 64) == 64
    split_values,
